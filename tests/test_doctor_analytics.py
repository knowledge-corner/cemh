"""
The doctor's Analytics tab: a filtered, paginated records table and a CSV
download, for a doctor researching her own patient population by condition,
age or gender.

Scoped to the doctor's own patients always — there is no Doctor or
Specialisation picker any more, since a doctor's analytics has nothing else
to be about.

The dashboard (stat tiles/charts) is disabled for now — the page always
shows the records list, unpaged sub-tab UI included, 10 patients per page.
"""

from datetime import timedelta
from io import BytesIO

from django.test import TestCase
from django.urls import reverse
from django.utils import timezone
from openpyxl import load_workbook

from accounts.models import DoctorProfile, Specialisation
from audit.models import AccessLog, AuditAction
from clinical.models import ClinicalNote, Diagnosis, Investigation
from patients.models import Sex
from pharmacy.models import Prescription, PrescriptionItem

from .factories import make_doctor, make_patient, make_receptionist, make_visit, today_at


class AnalyticsTestCase(TestCase):
    def setUp(self):
        self.doctor = make_doctor()
        self.receptionist = make_receptionist()
        self.client.force_login(self.doctor)

    def _list(self, **params):
        return self.client.get(reverse("doctor_analytics"), params)

    def _export(self, **params):
        return self.client.get(reverse("doctor_analytics_export"), params)


class TestAccessToAnalytics(AnalyticsTestCase):
    def test_a_doctor_can_open_it(self):
        self.assertEqual(self._list().status_code, 200)

    def test_a_receptionist_cannot(self):
        self.client.force_login(self.receptionist)
        self.assertEqual(self._list().status_code, 403)

    def test_a_receptionist_cannot_export_either(self):
        self.client.force_login(self.receptionist)
        self.assertEqual(self._export().status_code, 403)

    def test_the_nav_bar_offers_analytics(self):
        response = self.client.get(reverse("doctor_home"))
        self.assertContains(response, reverse("doctor_analytics"))

    def test_the_dashboard_tab_switcher_is_gone(self):
        response = self._list()
        self.assertNotContains(response, ">Dashboard<")
        self.assertNotContains(response, ">View<")


class TestTheUnfilteredView(AnalyticsTestCase):
    def setUp(self):
        super().setUp()
        aarav = make_patient(first_name="Aarav", phone="9820000101")
        bina = make_patient(first_name="Bina", phone="9820000102", sex=Sex.FEMALE)
        make_visit(aarav, self.doctor, start=today_at(9))
        make_visit(bina, self.doctor, start=today_at(10))

    def test_lists_every_patient_with_no_filters(self):
        response = self._list()
        self.assertEqual(response.context["download_count"], 2)
        self.assertEqual(len(response.context["rows"]), 2)


class TestPagination(AnalyticsTestCase):
    def setUp(self):
        super().setUp()
        for i in range(23):
            patient = make_patient(first_name=f"Pager{i:02d}", phone=f"98200003{i:02d}")
            make_visit(patient, self.doctor, start=today_at(9) + timedelta(minutes=30 * i))

    def test_the_first_page_shows_ten(self):
        response = self._list()
        self.assertEqual(len(response.context["rows"]), 10)
        self.assertEqual(response.context["page_obj"].paginator.num_pages, 3)
        self.assertEqual(response.context["download_count"], 23)

    def test_the_last_page_shows_the_remainder(self):
        response = self._list(page=3)
        self.assertEqual(len(response.context["rows"]), 3)

    def test_pagination_links_appear_when_there_is_more_than_one_page(self):
        response = self._list()
        self.assertContains(response, "Page 1 of 3")
        self.assertContains(response, "Next")

    def test_an_out_of_range_page_falls_back_to_the_last_one(self):
        response = self._list(page=999)
        self.assertEqual(response.context["page_obj"].number, 3)


class TestGenderAndAgeFilters(AnalyticsTestCase):
    def setUp(self):
        super().setUp()
        self.child = make_patient(
            first_name="ChildOne", phone="9820000201", sex=Sex.MALE,
            date_of_birth=timezone.localdate() - timedelta(days=int(8 * 365.25)),
        )
        self.adult = make_patient(
            first_name="AdultOne", phone="9820000202", sex=Sex.FEMALE,
            date_of_birth=timezone.localdate() - timedelta(days=int(45 * 365.25)),
        )
        make_visit(self.child, self.doctor, start=today_at(9))
        make_visit(self.adult, self.doctor, start=today_at(10))

    def test_filtering_by_gender(self):
        response = self._list(sex=Sex.FEMALE)
        self.assertEqual(response.context["download_count"], 1)

    def test_filtering_by_minimum_age_excludes_the_child(self):
        response = self._list(age_min=18)
        self.assertEqual(response.context["download_count"], 1)

    def test_filtering_by_maximum_age_excludes_the_adult(self):
        response = self._list(age_max=17)
        self.assertEqual(response.context["download_count"], 1)

    def test_an_age_band_can_exclude_everybody(self):
        response = self._list(age_min=20, age_max=30)
        self.assertEqual(response.context["download_count"], 0)


class TestConditionAndStatusFilters(AnalyticsTestCase):
    def setUp(self):
        super().setUp()
        self.thyroid_patient = make_patient(first_name="Thy", phone="9820000301")
        Diagnosis.objects.create(
            patient=self.thyroid_patient, description="Hypothyroidism",
            status=Diagnosis.Status.ACTIVE,
        )
        self.mixed_patient = make_patient(first_name="Mix", phone="9820000302")
        Diagnosis.objects.create(
            patient=self.mixed_patient, description="Hypothyroidism",
            status=Diagnosis.Status.ACTIVE,
        )
        Diagnosis.objects.create(
            patient=self.mixed_patient, description="Diabetes",
            status=Diagnosis.Status.RESOLVED,
        )
        self.unrelated_patient = make_patient(first_name="Un", phone="9820000303")
        Diagnosis.objects.create(
            patient=self.unrelated_patient, description="Migraine",
            status=Diagnosis.Status.ACTIVE,
        )
        for i, patient in enumerate(
            (self.thyroid_patient, self.mixed_patient, self.unrelated_patient)
        ):
            make_visit(patient, self.doctor, start=today_at(9 + i))

    def test_matching_by_condition_text_is_case_insensitive(self):
        response = self._list(condition="THYROID")
        self.assertEqual(response.context["download_count"], 2)

    def test_combining_condition_and_status_matches_the_same_diagnosis(self):
        # mixed_patient's Hypothyroidism is ACTIVE and their Diabetes is
        # RESOLVED — no single diagnosis is both "thyroid" and RESOLVED, so
        # this must not match, even though each half matches something.
        response = self._list(condition="thyroid", diagnosis_status=Diagnosis.Status.RESOLVED)
        self.assertEqual(response.context["download_count"], 0)

    def test_combining_condition_and_status_when_they_do_agree(self):
        response = self._list(condition="thyroid", diagnosis_status=Diagnosis.Status.ACTIVE)
        self.assertEqual(response.context["download_count"], 2)


class TestAnalyticsIsScopedToOwnPatients(AnalyticsTestCase):
    """
    A doctor's analytics is implicitly "my own patients" — there is no
    Doctor or Specialisation picker to narrow it with any more, and a patient
    who has only ever been seen by somebody else does not appear, whatever a
    crafted querystring claims.
    """

    def setUp(self):
        super().setUp()
        self.endo = Specialisation.objects.create(name="Test Endocrine Research")
        self.cardio = Specialisation.objects.create(name="Test Cardiology Research")
        DoctorProfile.objects.create(user=self.doctor, specialisation=self.endo)

        self.other_doctor = make_doctor(username="drother", email="other@example.in")
        DoctorProfile.objects.create(user=self.other_doctor, specialisation=self.cardio)

        self.my_patient = make_patient(first_name="Mine", phone="9820000401")
        make_visit(self.my_patient, self.doctor, start=today_at(10))

        self.their_patient = make_patient(first_name="Theirs", phone="9820000402")
        make_visit(self.their_patient, self.other_doctor, start=today_at(11))

    def test_only_my_own_patient_is_counted(self):
        response = self._list()
        self.assertEqual(response.context["download_count"], 1)

    def test_the_doctor_and_specialisation_filters_are_not_offered(self):
        response = self._list()
        self.assertNotContains(response, 'name="doctor"')
        self.assertNotContains(response, 'name="specialisation"')

    def test_a_crafted_doctor_parameter_is_ignored(self):
        # Even asked directly for the other doctor's patients, this doctor's
        # own analytics never shows anyone but their own.
        response = self._list(doctor=self.other_doctor.pk)
        self.assertEqual(response.context["download_count"], 1)
        self.assertEqual(response.context["filters"]["doctor"], self.doctor.pk)

    def test_it_stays_scoped(self):
        response = self._list(doctor=self.other_doctor.pk)
        self.assertEqual(response.context["download_count"], 1)
        self.assertEqual(response.context["rows"][0]["patient"], self.my_patient)


class TestDateRangeFilter(AnalyticsTestCase):
    def setUp(self):
        super().setUp()
        self.recent_patient = make_patient(first_name="Recent", phone="9820000501")
        make_visit(self.recent_patient, self.doctor, start=today_at(9))

        self.old_patient = make_patient(first_name="Old", phone="9820000502")
        make_visit(
            self.old_patient, self.doctor,
            start=timezone.now() - timedelta(days=400),
        )

    def test_a_date_range_excludes_visits_outside_it(self):
        today = timezone.localdate()
        response = self._list(
            date_from=(today - timedelta(days=1)).isoformat(),
            date_to=(today + timedelta(days=1)).isoformat(),
        )
        self.assertEqual(response.context["download_count"], 1)


class TestTheExcelExport(AnalyticsTestCase):
    """
    The "Download" button is a local backup, not the CSV list export: one
    workbook, one sheet per kind of clinical record, booking history left
    out since a backup for a doctor's own reference has no use for it.
    """

    def setUp(self):
        super().setUp()
        self.patient = make_patient(first_name="Downloadable", phone="9820000701")
        Diagnosis.objects.create(
            patient=self.patient, description="Hypothyroidism",
            status=Diagnosis.Status.ACTIVE,
        )
        excluded = make_patient(first_name="ExcludedByFilter", phone="9820000702")
        self.visit = make_visit(self.patient, self.doctor, start=today_at(9))
        make_visit(excluded, self.doctor, start=today_at(10))

    def _workbook(self, **params):
        response = self._export(**params)
        return response, load_workbook(BytesIO(response.content))

    def test_it_downloads_an_excel_workbook(self):
        response = self._export()
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.assertIn("attachment", response["Content-Disposition"])
        self.assertIn(".xlsx", response["Content-Disposition"])

    def test_it_has_one_sheet_per_kind_of_record(self):
        _, wb = self._workbook()
        self.assertEqual(
            wb.sheetnames, ["Patients", "Clinical Notes", "Investigations", "Prescriptions"]
        )

    def test_booking_history_is_not_a_sheet(self):
        _, wb = self._workbook()
        self.assertNotIn("Bookings", wb.sheetnames)
        self.assertNotIn("Visits", wb.sheetnames)

    def test_only_the_filtered_patients_are_rows(self):
        _, wb = self._workbook(condition="thyroid")
        names = [row[1].value for row in wb["Patients"].iter_rows(min_row=2)]
        self.assertIn("Downloadable", names)
        self.assertNotIn("ExcludedByFilter", names)

    def test_the_patients_sheet_header_names_the_columns(self):
        _, wb = self._workbook()
        header = [cell.value for cell in next(wb["Patients"].iter_rows(max_row=1))]
        self.assertIn("Gender", header)
        self.assertIn("Active diagnoses", header)

    def test_a_clinical_note_appears_on_its_own_sheet(self):
        ClinicalNote.objects.create(
            visit=self.visit, patient=self.patient, author=self.doctor,
            prescription_note="Take rest.",
        )
        _, wb = self._workbook()
        body = [[c.value for c in row] for row in wb["Clinical Notes"].iter_rows(min_row=2)]
        self.assertTrue(any("Take rest." in (row[-1] or "") for row in body))

    def test_an_investigation_appears_on_its_own_sheet(self):
        Investigation.objects.create(
            patient=self.patient, visit=self.visit, test_name="TSH", value="2.1",
        )
        _, wb = self._workbook()
        body = [[c.value for c in row] for row in wb["Investigations"].iter_rows(min_row=2)]
        self.assertTrue(any(row[3] == "TSH" for row in body))

    def test_a_prescription_item_appears_on_its_own_sheet(self):
        prescription = Prescription.objects.create(
            visit=self.visit, patient=self.patient, doctor=self.doctor,
        )
        PrescriptionItem.objects.create(prescription=prescription, drug_name="Levothyroxine")
        _, wb = self._workbook()
        body = [[c.value for c in row] for row in wb["Prescriptions"].iter_rows(min_row=2)]
        self.assertTrue(any(row[4] == "Levothyroxine" for row in body))

    def test_date_range_trims_records_outside_it(self):
        old_note = ClinicalNote.objects.create(
            visit=self.visit, patient=self.patient, author=self.doctor,
            prescription_note="Old note.",
        )
        ClinicalNote.objects.filter(pk=old_note.pk).update(
            created_at=timezone.now() - timedelta(days=400)
        )
        today = timezone.localdate()
        _, wb = self._workbook(
            date_from=(today - timedelta(days=1)).isoformat(),
            date_to=(today + timedelta(days=1)).isoformat(),
        )
        body = [[c.value for c in row] for row in wb["Clinical Notes"].iter_rows(min_row=2)]
        self.assertFalse(any("Old note." in (row[-1] or "") for row in body))

    def test_exporting_is_audited(self):
        self._export()
        entry = AccessLog.objects.filter(action=AuditAction.PRINT).get()
        self.assertEqual(entry.username, self.doctor.username)
